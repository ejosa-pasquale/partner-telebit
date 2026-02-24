import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl


APP_TITLE = "EV Field Service – Pricing & Margini (Partner vs Cliente)"
DATA_DIR = Path("data")
PARTNER_DIR = DATA_DIR / "partners"
DEFAULT_DIR = DATA_DIR / "defaults"

PARTNER_DIR.mkdir(parents=True, exist_ok=True)
DEFAULT_DIR.mkdir(parents=True, exist_ok=True)

# Precarico opzionale del listino cliente (metti qui il file nel repo)
DEFAULT_CLIENT_XLSX = DEFAULT_DIR / "client_pricelist.xlsx"

ITEM_RE = re.compile(r"^\s*Item\s*([0-9]+(?:\.[a-zA-Z])?)\s*[:\-]?\s*(.*)$")

@st.cache_data(show_spinner=False)
def parse_pricing_matrix_xlsx_cached(file_bytes: bytes) -> pd.DataFrame:
    return parse_pricing_matrix_xlsx(file_bytes)

def parse_pricing_matrix_xlsx(file_bytes: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    max_row, max_col = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]

    rows_out = []
    r = 0
    current_block = None
    distances = None  # list of (col_idx, distance_label)

    while r < len(grid):
        row = grid[r]
        b = row[1] if len(row) > 1 else None  # column B

        if isinstance(b, str) and b.strip().lower().startswith("installazione"):
            current_block = b.strip()
            distances = []
            for ci in range(2, len(row)):  # columns C...
                v = row[ci]
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                distances.append((ci, str(v).strip()))
            r += 1
            continue

        if current_block and distances:
            activity = b
            if isinstance(activity, str):
                m = ITEM_RE.match(activity.strip())
                if m:
                    item_id = m.group(1).strip()
                    item_desc = m.group(2).strip() if m.group(2) else ""
                    for ci, dist_label in distances:
                        price = row[ci]
                        if price is None or price == "":
                            continue
                        try:
                            price = float(price)
                        except Exception:
                            continue
                        rows_out.append(
                            {
                                "block": current_block,
                                "distance": dist_label,
                                "item_id": item_id,
                                "item_desc": item_desc,
                                "full_activity": activity.strip(),
                                "price": price,
                            }
                        )

        if current_block and all((v is None or (isinstance(v, str) and v.strip() == "")) for v in row):
            current_block = None
            distances = None

        r += 1

    df = pd.DataFrame(rows_out)
    if df.empty:
        raise ValueError("Non riesco a leggere la matrice: controlla che sia nello stesso formato del template.")
    return df


def save_upload(bytes_data: bytes, dst: Path):
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(bytes_data)


def load_xlsx_from_path(path: Path) -> bytes:
    return path.read_bytes()


def format_eur(x: float) -> str:
    return f"€ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def load_total_override(override_csv_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(override_csv_bytes))
    df.columns = [c.strip() for c in df.columns]
    required = {"block", "distance", "partner_total_override"}
    if not required.issubset(set(df.columns)):
        raise ValueError(f"Colonne richieste: {', '.join(sorted(required))}. Opzionale: region")
    df["partner_total_override"] = pd.to_numeric(df["partner_total_override"], errors="coerce")
    df = df.dropna(subset=["partner_total_override"])
    df["block"] = df["block"].astype(str).str.strip()
    df["distance"] = df["distance"].astype(str).str.strip()
    if "region" in df.columns:
        df["region"] = df["region"].astype(str).str.strip()
    return df


def get_override_total_value(df_override: pd.DataFrame, region: str, block: str, distance: str):
    if df_override is None or df_override.empty:
        return None
    block = str(block).strip()
    distance = str(distance).strip()
    if "region" in df_override.columns:
        region = str(region).strip()
        hit = df_override[(df_override["region"] == region) & (df_override["block"] == block) & (df_override["distance"] == distance)]
    else:
        hit = df_override[(df_override["block"] == block) & (df_override["distance"] == distance)]
    if hit.empty:
        return None
    return float(hit.iloc[0]["partner_total_override"])


st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)
st.caption("Prezzario cliente (anche precaricato) + listini partner per regione. Margine su N installazioni + rebate 5%.")

with st.sidebar:
    st.header("Configurazione")
    rebate_pct = st.number_input("Rebate al cliente finale (%)", min_value=0.0, max_value=100.0, value=5.0, step=0.5) / 100.0

    st.markdown("---")
    st.subheader("Prezzario Cliente")
    has_default_client = DEFAULT_CLIENT_XLSX.exists()
    client_mode = st.radio(
        "Sorgente prezzario cliente",
        options=(["Usa precaricato (repo)"] if has_default_client else []) + ["Carica file Excel"],
        index=0 if has_default_client else 0,
    )

    client_file_bytes = None
    if client_mode.startswith("Usa precaricato"):
        client_file_bytes = load_xlsx_from_path(DEFAULT_CLIENT_XLSX)
        st.success(f"Usando prezzario precaricato: {DEFAULT_CLIENT_XLSX.name}")
        with st.expander("Sostituire il prezzario precaricato"):
            st.markdown(f"Carica il tuo file nel repo in `{DEFAULT_CLIENT_XLSX.as_posix()}` con nome `{DEFAULT_CLIENT_XLSX.name}`.")
    else:
        client_upl = st.file_uploader("Carica prezzario cliente (xlsx)", type=["xlsx"], key="client_upl")
        if client_upl:
            client_file_bytes = client_upl.getvalue()

    st.markdown("---")
    st.subheader("Listini Partner (persistenti)")
    st.caption("Carica una volta per regione: viene salvato su disco (cartella data/partners).")

    region_input = st.text_input("Regione (es. Lombardia, Lazio, ...)", value="")
    partner_file = st.file_uploader("File Excel listino partner (xlsx)", type=["xlsx"], key="partner_upl")
    if st.button("Salva listino partner", disabled=not (region_input.strip() and partner_file)):
        save_upload(partner_file.getvalue(), PARTNER_DIR / f"{region_input.strip()}.xlsx")
        st.success(f"Salvato: {region_input.strip()}.xlsx")

    st.markdown("**Regioni salvate**")
    partner_files = sorted([p for p in PARTNER_DIR.glob("*.xlsx")])
    if partner_files:
        for p in partner_files:
            c1, c2 = st.columns([3, 1])
            c1.write(p.stem)
            if c2.button("Elimina", key=f"del_{p.stem}"):
                p.unlink(missing_ok=True)
                st.rerun()
    else:
        st.info("Nessun listino partner salvato ancora.")

    st.markdown("---")
    st.subheader("Override totale partner (opzionale)")
    st.caption("CSV: block,distance,partner_total_override (opzionale: region)")
    override_total_csv = st.file_uploader("Override Totale (CSV)", type=["csv"], key="override_total_upl")

partner_files = sorted([p for p in PARTNER_DIR.glob("*.xlsx")])
regions_available = [p.stem for p in partner_files]

if client_file_bytes is None:
    st.info("Carica (o precarica) il prezzario Cliente per iniziare.")
    st.stop()

if not regions_available:
    st.warning("Carica almeno un listino partner in sidebar (una regione).")
    st.stop()

override_total_df = None
if override_total_csv:
    try:
        override_total_df = load_total_override(override_total_csv.getvalue())
        st.sidebar.success("Override totale caricato.")
    except Exception as e:
        st.sidebar.error(f"Override totale non valido: {e}")

colA, colB = st.columns([1, 2], gap="large")
with colA:
    st.subheader("Selezione")
    selected_region = st.selectbox("Regione (partner)", options=regions_available)
    qty_install = st.number_input("Numero installazioni", min_value=1, value=1, step=1)

with colB:
    st.subheader("Input operativi")
    st.markdown("Scegli il pacchetto (tipo installazione + distanza) e quali Item includere.")

try:
    df_client = parse_pricing_matrix_xlsx_cached(client_file_bytes).rename(columns={"price": "client_price"})
except Exception as e:
    st.error(f"Errore parsing prezzario cliente: {e}")
    st.stop()

partner_path = PARTNER_DIR / f"{selected_region}.xlsx"
try:
    df_partner = parse_pricing_matrix_xlsx_cached(load_xlsx_from_path(partner_path)).rename(columns={"price": "partner_price"})
except Exception as e:
    st.error(f"Errore parsing listino partner ({selected_region}): {e}")
    st.stop()

key_cols = ["block", "distance", "item_id"]
df = df_client.merge(df_partner[key_cols + ["partner_price"]], on=key_cols, how="left", validate="m:1")

missing_partner = df["partner_price"].isna().sum()
if missing_partner:
    st.warning(f"Attenzione: {missing_partner} righe del cliente non hanno corrispondenza nel listino partner della regione selezionata.")

blocks = sorted(df["block"].unique())
sel_block = st.selectbox("Tipo installazione", options=blocks)
distances = sorted(df.loc[df["block"] == sel_block, "distance"].unique())
sel_dist = st.selectbox("Distanza dal contatore", options=distances)

df_sel = df[(df["block"] == sel_block) & (df["distance"] == sel_dist)].copy()
df_sel["include"] = True
df_sel = df_sel.sort_values(by=["item_id"])

st.markdown("#### Item inclusi")
edited = st.data_editor(
    df_sel[["include", "item_id", "full_activity", "client_price", "partner_price"]],
    use_container_width=True,
    disabled=["item_id", "full_activity", "client_price", "partner_price"],
    column_config={
        "client_price": st.column_config.NumberColumn("Cliente (€)", format="%.2f"),
        "partner_price": st.column_config.NumberColumn("Partner (€)", format="%.2f"),
        "full_activity": st.column_config.TextColumn("Attività / Item"),
    },
    hide_index=True,
)

df_sel["include"] = edited["include"]
included = df_sel[df_sel["include"]].copy()
if included.empty:
    st.warning("Seleziona almeno un item da includere.")
    st.stop()

client_total_unit = float(included["client_price"].sum())
partner_total_unit_calc = float(included["partner_price"].fillna(0).sum())

override_total_value = get_override_total_value(override_total_df, selected_region, sel_block, sel_dist)
used_override_total = override_total_value is not None
partner_total_unit = float(override_total_value) if used_override_total else partner_total_unit_calc

margin_unit = client_total_unit - partner_total_unit
gross_margin = margin_unit * float(qty_install)
rebate = gross_margin * rebate_pct
net_profit = gross_margin - rebate

k1, k2, k3, k4 = st.columns(4)
k1.metric("Totale Cliente (unitario)", format_eur(client_total_unit))
k2.metric("Totale Partner (unitario)", format_eur(partner_total_unit))
k3.metric("Margine lordo totale", format_eur(gross_margin))
k4.metric(f"Guadagno netto (rebate {rebate_pct*100:.1f}%)", format_eur(net_profit))

if used_override_total:
    st.info("✅ Override totale partner applicato per questo pacchetto (tipo installazione + distanza).")
else:
    st.caption("Override totale partner non presente: totale partner = somma degli item partner selezionati.")

if margin_unit < 0:
    st.error("Margine unitario negativo: totale partner > totale cliente. Controlla listini o override.")

st.markdown("#### Dettaglio item (per trasparenza)")
detail = included[["item_id", "full_activity", "client_price", "partner_price"]].copy()
st.dataframe(detail, use_container_width=True)

st.markdown("#### Esporta report")
summary = pd.DataFrame([{
    "regione": selected_region,
    "tipo_installazione": sel_block,
    "distanza": sel_dist,
    "numero_installazioni": qty_install,
    "totale_cliente_unit": client_total_unit,
    "totale_partner_unit": partner_total_unit,
    "override_totale_partner_usato": used_override_total,
    "override_totale_partner_valore": override_total_value if used_override_total else None,
    "margine_unit": margin_unit,
    "margine_lordo_totale": gross_margin,
    "rebate_pct": rebate_pct,
    "rebate": rebate,
    "guadagno_netto": net_profit,
}])

out = io.BytesIO()
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    summary.to_excel(writer, index=False, sheet_name="Summary")
    detail.to_excel(writer, index=False, sheet_name="Dettaglio_Item")

st.download_button(
    "Scarica report Excel",
    data=out.getvalue(),
    file_name=f"report_{selected_region}_{sel_block[:15].replace(' ','_')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown("---")
with st.expander("Formato Override Totale (CSV)"):
    st.markdown(
        """CSV richiesto:
- block
- distance
- partner_total_override
Opzionale:
- region

Esempio:
```csv
region,block,distance,partner_total_override
Lombardia,Installazione Wallbox 7,4 kW monofase,2 mt. dal contatore,520
```
"""
    )
